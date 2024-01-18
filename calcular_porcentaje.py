import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment,Font
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
def calcularPorcent(data,porcent,final_path,condition,cal):

    df = pd.DataFrame(data)
    # procesamos el dataframe
    valores_unicos_lista = df['YEAR'].unique().tolist()
    meses_unicos = []
    for year in valores_unicos_lista:
        c1 = df['YEAR'] == year
        datos_seleccionados = df[c1]
        meses_unicos.append({'meses':datos_seleccionados['MES'].unique().tolist(),'year':year})
    longitud_meses = 0
    for lon in meses_unicos:
        longitud_meses += len(lon['meses'])
    size = df.shape[0]
    total_porcent = 24
    muestra_generada = []
    if condition == 1:
        total_porcent = round(int(size)*1)
        muestra_generada = data
    elif condition == 2:
        N_poblation = size
        error = 0.05 
        confianza = 1.96 
        total_porcent = (N_poblation *(confianza**2)*0.5*0.5)/((error**2)*(N_poblation-1)+(confianza**2)*0.5*0.5)
      
        for dic in meses_unicos:
            c2 = df['YEAR'] == dic['year']
            data_frame_year = df[c2]
            for val in dic['meses']:
                c3 = data_frame_year['MES'] == val
                df_mes = data_frame_year[c3]
                elementos_aleatorios = df_mes.sample(n=ctd_mes, replace=False)
                aux = elementos_aleatorios.to_dict(orient='records')
                muestra_generada.extend(aux)
    elif condition == 3:
        total_porcent = round(int(size)*porcent)
        ctd_mes = round(total_porcent/longitud_meses)
        for dic in meses_unicos:
            c2 = df['YEAR'] == dic['year']
            data_frame_year = df[c2]
            for val in dic['meses']:
                c3 = data_frame_year['MES'] == val
                df_mes = data_frame_year[c3]
                elementos_aleatorios = df_mes.sample(n=ctd_mes, replace=False)
                aux = elementos_aleatorios.to_dict(orient='records')
                muestra_generada.extend(aux)
    

    if cal == 20:
        workbook = load_workbook('./plantillas/cal20.xlsx')
        sheet = workbook.active
        aux_init = 3
        
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['YEAR']
            sheet[f'B{aux_init}'] = i['MES']
            sheet[f'C{aux_init}'] = i['DIA']
            sheet[f'D{aux_init}'] = i['Subestación / Barra']
            sheet[f'E{aux_init}'] = i['GEO-X']
            sheet[f'F{aux_init}'] = i['GEO-Y']
            sheet[f'G{aux_init}'] = i['Provincia']
            sheet[f'H{aux_init}'] = i['Cantón']
            sheet[f'K{aux_init}'] = i['# Registros']
            sheet[f'L{aux_init}'] = i['Fase A V']
            sheet[f'M{aux_init}'] = i['Fase B V']
            sheet[f'N{aux_init}'] = i['Fase C V']
            sheet[f'T{aux_init}'] = i['OBSERVACIONES']
            aux_init += 1
        
        workbook.save(final_path)
        workbook.close()
    elif cal == 30:
        workbook = load_workbook('./plantillas/cal30.xlsx')
        sheet = workbook.active
        aux_init = 4
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['YEAR']
            sheet[f'B{aux_init}'] = i['MES']
            sheet[f'C{aux_init}'] = i['DIA']
            sheet[f'D{aux_init}'] = i['CODIGO']
            sheet[f'E{aux_init}'] = i['TIPO']
            sheet[f'F{aux_init}'] = i['SUBESTACION']
            sheet[f'G{aux_init}'] = i['ALIMENTADOR']
            sheet[f'H{aux_init}'] = i['NUM DE FASES']
            sheet[f'I{aux_init}'] = i['F-F']
            sheet[f'J{aux_init}'] = i['F-N']
            sheet[f'L{aux_init}'] = i['N_REGISTROS']
            sheet[f'M{aux_init}'] = i['FA_V']
            sheet[f'N{aux_init}'] = i['FA_PST']
            sheet[f'L{aux_init}'] = i['N_REGISTROS']
            sheet[f'M{aux_init}'] = i['FA_V']
            sheet[f'N{aux_init}'] = i['FA_PST']
            sheet[f'O{aux_init}'] = i['FA_VTHD']
            sheet[f'P{aux_init}'] = i['FC_V']
            sheet[f'Q{aux_init}'] = i['FC_PST']
            sheet[f'R{aux_init}'] = i['FC_VTHD']
            sheet[f'S{aux_init}'] = i['FB_V']
            sheet[f'T{aux_init}'] = i['FB_PST']
            sheet[f'U{aux_init}'] = i['FB_VTHD']
            sheet[f'V{aux_init}'] = i['DESEQUILIBRIO']
            sheet[f'AI{aux_init}'] = i['OBSERVACIONES']
            aux_init += 1
        workbook.save(final_path)
        workbook.close()
    elif cal == 40:
        workbook = load_workbook('./plantillas/cal40.xlsx')
        sheet = workbook.active
        aux_init = 3
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['YEAR']
            sheet[f'B{aux_init}'] = i['MES']
            sheet[f'C{aux_init}'] = i['DIA']
            sheet[f'D{aux_init}'] = i['CODIGO_UNICO_NACIONAL']
            sheet[f'E{aux_init}'] = i['TIPO']
            sheet[f'F{aux_init}'] = i['PROVINCIA']
            sheet[f'G{aux_init}'] = i['CANTON']
            sheet[f'H{aux_init}'] = i['SUBESTACION']

            aux_init += 1
        workbook.save(final_path)
        workbook.close()
    #df = pd.DataFrame(muestra_generada)
    #df.to_excel(final_path, index=False)



