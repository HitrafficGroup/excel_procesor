import openpyxl
import string
path_file = 'C:\Users\david\Desktop\python_excel\CAL 020 v19.1 Barras EEASA Enero 2021.xlsx'
workbook = openpyxl.load_workbook(path_file)
lista_de_hojas = workbook.sheetnames
datos_hoja = []

#abrimos la hoja 1
for name in lista_de_hojas:
    hoja = workbook[name]
    for fila in hoja.iter_rows(min_row=12, values_only=True):
        # Suponiendo que la primera fila tiene encabezados y se desea incluirlos en el diccionario
        encabezados = [celda.value for celda in hoja[1]]
        fila_dict = dict(zip(encabezados, fila))
        datos_hoja.append(fila_dict)
nombres_unicos = []

for i in datos_hoja:
    lista_de_claves = list(i) 
