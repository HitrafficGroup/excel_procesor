import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment,Font
from openpyxl.drawing.image import Image
def calcularPorcent(data,porcent,final_path,condition,cal):

    df = pd.DataFrame(data)
    # procesamos el dataframe
    valores_unicos_lista = df['YEAR'].unique().tolist()
    meses_unicos = []
    for year in valores_unicos_lista:
        c1 = df['YEAR'] == year
        datos_seleccionados = df[c1]
        meses_unicos.append({'meses':datos_seleccionados['MES'].unique().tolist(),'year':year})
    longitud_meses = 0
    for lon in meses_unicos:
        longitud_meses += len(lon['meses'])
    size = df.shape[0]
    total_porcent = 24
    muestra_generada = []
    if condition == 1:
        total_porcent = round(int(size)*1)
        muestra_generada = data
    elif condition == 2:
        N_poblation = size
        error = 0.05 
        confianza = 1.96 
        total_porcent = (N_poblation *(confianza**2)*0.5*0.5)/((error**2)*(N_poblation-1)+(confianza**2)*0.5*0.5)
      
        for dic in meses_unicos:
            c2 = df['YEAR'] == dic['year']
            data_frame_year = df[c2]
            for val in dic['meses']:
                c3 = data_frame_year['MES'] == val
                df_mes = data_frame_year[c3]
                elementos_aleatorios = df_mes.sample(n=ctd_mes, replace=False)
                aux = elementos_aleatorios.to_dict(orient='records')
                muestra_generada.extend(aux)
    elif condition == 3:
        total_porcent = round(int(size)*porcent)
        ctd_mes = round(total_porcent/longitud_meses)
        for dic in meses_unicos:
            c2 = df['YEAR'] == dic['year']
            data_frame_year = df[c2]
            for val in dic['meses']:
                c3 = data_frame_year['MES'] == val
                df_mes = data_frame_year[c3]
                elementos_aleatorios = df_mes.sample(n=ctd_mes, replace=False)
                aux = elementos_aleatorios.to_dict(orient='records')
                muestra_generada.extend(aux)
    

    if cal == 20:
        workbook = openpyxl.Workbook()
        borde_negro_grueso = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
        centrar_texto = Alignment(horizontal='center', vertical='center')
        sheet = workbook.active
        sheet.column_dimensions['D'].width = 14
        sheet.column_dimensions['G'].width = 11
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 15
        sheet.column_dimensions['O'].width = 15
        sheet.column_dimensions['S'].width = 30
        sheet.column_dimensions['T'].width = 30
        sheet['A1'].border = borde_negro_grueso
        sheet['A1'].alignment = centrar_texto
        sheet['A1'].font = Font(bold=True)
        sheet['A1'] = 'YEAR'
        sheet.merge_cells('A1:A2')
        sheet['B1'].border = borde_negro_grueso
        sheet['B1'].alignment = centrar_texto
        sheet['B1'].font = Font(bold=True)
        sheet['B1'] = 'MES'
        sheet.merge_cells('B1:B2')
        sheet['C1'].border = borde_negro_grueso
        sheet['C1'].alignment = centrar_texto
        sheet['C1'].font = Font(bold=True)
        sheet['C1'] = 'DIA'
        sheet.merge_cells('C1:C2')
        sheet['D1'].border = borde_negro_grueso
        sheet['D1'].alignment = centrar_texto
        sheet['D1'].font = Font(bold=True)
        sheet['D1'] = 'SUBESTACION'
        sheet.merge_cells('D1:D2')
        sheet['E1'].border = borde_negro_grueso
        sheet['E1'].alignment = centrar_texto
        sheet['E1'].font = Font(bold=True)
        sheet['E1'] = 'GEO-X'
        sheet.merge_cells('E1:E2')
        sheet['F1'].border = borde_negro_grueso
        sheet['F1'].alignment = centrar_texto
        sheet['F1'].font = Font(bold=True)
        sheet['F1'] = 'GEO-Y'
        sheet.merge_cells('F1:F2')
        sheet['G1'].border = borde_negro_grueso
        sheet['G1'].alignment = centrar_texto
        sheet['G1'].font = Font(bold=True)
        sheet['G1'] = 'PROVINCIA'
        sheet.merge_cells('G1:G2')
        sheet['H1'].border = borde_negro_grueso
        sheet['H1'].alignment = centrar_texto
        sheet['H1'].font = Font(bold=True)
        sheet['H1'] = 'CANTON'
        sheet.merge_cells('H1:H2')
        sheet['I1'].border = borde_negro_grueso
        sheet['I1'].alignment = Alignment(wrap_text=True)
        sheet['I1'].font = Font(bold=True)
        sheet['I1'] = 'ARCHIVO FUENTE'
        sheet.merge_cells('I1:I2')
        sheet['J1'].border = borde_negro_grueso
        sheet['J1'].alignment = Alignment(wrap_text=True)
        sheet['J1'].font = Font(bold=True)
        sheet['J1'] = 'SE PLANIFICO'
        sheet.merge_cells('J1:J2')
        sheet['K1'].border = borde_negro_grueso
        sheet['K1'].alignment = centrar_texto
        sheet['K1'].font = Font(bold=True)
        sheet['K1'] = 'REPORTADOS'
        sheet.merge_cells('K1:N1')
        sheet['K2'].border = borde_negro_grueso
        sheet['K2'].alignment = centrar_texto
        sheet['K2'].font = Font(bold=True)
        sheet['K2'] = '# Mediciones (8)'

        sheet['L2'].border = borde_negro_grueso
        sheet['L2'].alignment = centrar_texto
        sheet['L2'].font = Font(bold=True)
        sheet['L2'] = '∆V FA (39)'

        sheet['M2'].border = borde_negro_grueso
        sheet['M2'].alignment = centrar_texto
        sheet['M2'].font = Font(bold=True)
        sheet['M2'] = '∆V FB (52)'

        sheet['N2'].border = borde_negro_grueso
        sheet['N2'].alignment = centrar_texto
        sheet['N2'].font = Font(bold=True)
        sheet['N2'] = '∆V FC(63)'

        sheet['O1'].border = borde_negro_grueso
        sheet['O1'].alignment = centrar_texto
        sheet['O1'].font = Font(bold=True)
        sheet['O1'] = 'MEDICIONES Auditadas '
        sheet.merge_cells('O1:R1')

        sheet['O2'].border = borde_negro_grueso
        sheet['O2'].alignment = centrar_texto
        sheet['O2'].font = Font(bold=True)
        sheet['O2'] = '# Mediciones'

        sheet['P2'].border = borde_negro_grueso
        sheet['P2'].alignment = centrar_texto
        sheet['P2'].font = Font(bold=True)
        sheet['P2'] = '∆V FA'

        sheet['Q2'].border = borde_negro_grueso
        sheet['Q2'].alignment = centrar_texto
        sheet['Q2'].font = Font(bold=True)
        sheet['Q2'] = '∆V FB'

        sheet['R2'].border = borde_negro_grueso
        sheet['R2'].alignment = centrar_texto
        sheet['R2'].font = Font(bold=True)
        sheet['R2'] = '∆V FC'

        sheet['S1'].border = borde_negro_grueso
        sheet['S1'].alignment = centrar_texto
        sheet['S1'].font = Font(bold=True)
        sheet['S1'] = 'OBSERVACIONES AUDITADAS'
        sheet.merge_cells('S1:S2')

        sheet['T1'].border = borde_negro_grueso
        sheet['T1'].alignment = centrar_texto
        sheet['T1'].font = Font(bold=True)
        sheet['T1'] = 'OBSERVACIONES REPORTADAS'
        sheet.merge_cells('T1:T2')

        #a partir de aqui agregamos los datos
        aux_init = 3
        
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['YEAR']
            sheet[f'B{aux_init}'] = i['MES']
            sheet[f'C{aux_init}'] = i['DIA']
            sheet[f'D{aux_init}'] = i['Subestación / Barra']
            sheet[f'E{aux_init}'] = i['GEO-X']
            sheet[f'F{aux_init}'] = i['GEO-Y']
            sheet[f'G{aux_init}'] = i['Provincia']
            sheet[f'H{aux_init}'] = i['Cantón']
            sheet[f'K{aux_init}'] = i['# Registros']
            sheet[f'L{aux_init}'] = i['Fase A V']
            sheet[f'M{aux_init}'] = i['Fase B V']
            sheet[f'N{aux_init}'] = i['Fase C V']
            sheet[f'T{aux_init}'] = i['OBSERVACIONES']
            aux_init += 1
    
        workbook.save(final_path)
    if cal == 30:
        workbook = openpyxl.Workbook()
        borde_negro_grueso = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
        centrar_texto = Alignment(horizontal='center', vertical='center')
        sheet = workbook.active
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['L'].width = 16
        sheet.column_dimensions['V'].width = 16
        sheet.column_dimensions['W'].width = 16
        sheet['A1'].border = borde_negro_grueso
        sheet['A1'].alignment = centrar_texto
        sheet['A1'].font = Font(bold=True)
        sheet['A1'] = 'YEAR'
        sheet.merge_cells('A1:A3')
        sheet['B1'].border = borde_negro_grueso
        sheet['B1'].alignment = centrar_texto
        sheet['B1'].font = Font(bold=True)
        sheet['B1'] = 'MES'
        sheet.merge_cells('B1:B3')
        sheet['C1'].border = borde_negro_grueso
        sheet['C1'].alignment = centrar_texto
        sheet['C1'].font = Font(bold=True)
        sheet['C1'] = 'DIA'
        sheet.merge_cells('C1:C3')
        sheet['D1'].border = borde_negro_grueso
        sheet['D1'].alignment = centrar_texto
        sheet['D1'].font = Font(bold=True)
        sheet['D1'] = 'CODIGO'
        sheet.merge_cells('D1:D3')
        sheet['E1'].border = borde_negro_grueso
        sheet['E1'].alignment = centrar_texto
        sheet['E1'].font = Font(bold=True)
        sheet['E1'] = 'TIPO'
        sheet.merge_cells('E1:E3')
        sheet['F1'].border = borde_negro_grueso
        sheet['F1'].alignment = centrar_texto
        sheet['F1'].font = Font(bold=True)
        sheet['F1'] = 'SUBESTACION'
        sheet.merge_cells('F1:F3')
        sheet['G1'].border = borde_negro_grueso
        sheet['G1'].alignment = centrar_texto
        sheet['G1'].font = Font(bold=True)
        sheet['G1'] = 'ALIMENTADOR'
        sheet.merge_cells('G1:G3')
        sheet['H1'].border = borde_negro_grueso
        sheet['H1'].alignment = Alignment(wrap_text=True)
        sheet['H1'].font = Font(bold=True)
        sheet['H1'] = 'NUM DE FASES'
        sheet.merge_cells('H1:H3')
        sheet['I1'].border = borde_negro_grueso
        sheet['I1'].alignment = centrar_texto
        sheet['I1'].font = Font(bold=True)
        sheet['I1'] = 'F-F'
        sheet.merge_cells('I1:I3')
        sheet['J1'].border = borde_negro_grueso
        sheet['J1'].alignment = centrar_texto
        sheet['J1'].font = Font(bold=True)
        sheet['J1'] = 'F-N'
        sheet.merge_cells('J1:J3')
        sheet['K1'].border = borde_negro_grueso
        sheet['K1'].alignment = Alignment(wrap_text=True)
        sheet['K1'].font = Font(bold=True)
        sheet['K1'] = 'ARCHIVO FUENTE'
        sheet.merge_cells('K1:K3')
        
        sheet['L1'].border = borde_negro_grueso
        sheet['L1'].alignment = centrar_texto
        sheet['L1'].font = Font(bold=True)
        sheet['L1'] = 'REGISTROS FUERA DE LÍMITES REPORTADOS'
        sheet.merge_cells('L1:V1')

        sheet['L2'].border = borde_negro_grueso
        sheet['L2'].alignment = centrar_texto
        sheet['L2'].font = Font(bold=True)
        sheet['L2'] = '# Mediciones (18)'
        sheet.merge_cells('L2:L3')

        sheet['M2'].border = borde_negro_grueso
        sheet['M2'].alignment = centrar_texto
        sheet['M2'].font = Font(bold=True)
        sheet['M2'] = 'FASE A'
        sheet.merge_cells('M2:O2')

        sheet['M3'].border = borde_negro_grueso
        sheet['M3'].alignment = centrar_texto
        sheet['M3'].font = Font(bold=True)
        sheet['M3'] = 'V(19)'
      
        sheet['N3'].border = borde_negro_grueso
        sheet['N3'].alignment = centrar_texto
        sheet['N3'].font = Font(bold=True)
        sheet['N3'] = 'PST (20)'

        sheet['O3'].border = borde_negro_grueso
        sheet['O3'].alignment = centrar_texto
        sheet['O3'].font = Font(bold=True)
        sheet['O3'] = 'VTHD (21)'

        sheet['P3'].border = borde_negro_grueso
        sheet['P3'].alignment = centrar_texto
        sheet['P3'].font = Font(bold=True)
        sheet['P3'] = 'V (22)'


        sheet['Q3'].border = borde_negro_grueso
        sheet['Q3'].alignment = centrar_texto
        sheet['Q3'].font = Font(bold=True)
        sheet['Q3'] = 'PST (23)'
      
        sheet['R3'].border = borde_negro_grueso
        sheet['R3'].alignment = centrar_texto
        sheet['R3'].font = Font(bold=True)
        sheet['R3'] = 'VTHD (24)'

        sheet['S3'].border = borde_negro_grueso
        sheet['S3'].alignment = centrar_texto
        sheet['S3'].font = Font(bold=True)
        sheet['S3'] = 'V (25)'

        sheet['T3'].border = borde_negro_grueso
        sheet['T3'].alignment = centrar_texto
        sheet['T3'].font = Font(bold=True)
        sheet['T3'] = 'PST (26)'

        sheet['U3'].border = borde_negro_grueso
        sheet['U3'].alignment = centrar_texto
        sheet['U3'].font = Font(bold=True)
        sheet['U3'] = 'VTHD (27)'

        sheet['P2'].border = borde_negro_grueso
        sheet['P2'].alignment = centrar_texto
        sheet['P2'].font = Font(bold=True)
        sheet['P2'] = 'FASE B'
        sheet.merge_cells('P2:R2')

        sheet['S2'].border = borde_negro_grueso
        sheet['S2'].alignment = centrar_texto
        sheet['S2'].font = Font(bold=True)
        sheet['S2'] = 'FASE C'
        sheet.merge_cells('S2:U2')

        sheet['V2'].border = borde_negro_grueso
        sheet['V2'].alignment = centrar_texto
        sheet['V2'].font = Font(bold=True)
        sheet['V2'] = 'Desequilibrio (28)'
        sheet.merge_cells('V2:V3')

        #agregamos los datos de las columnas auditadas
        sheet['W1'].border = borde_negro_grueso
        sheet['W1'].alignment = centrar_texto
        sheet['W1'].font = Font(bold=True)
        sheet['W1'] = 'REGISTROS FUERA DE LÍMITES AUDITADOS'
        sheet.merge_cells('W1:AG1')

        sheet['W2'].border = borde_negro_grueso
        sheet['W2'].alignment = centrar_texto
        sheet['W2'].font = Font(bold=True)
        sheet['W2'] = '# Mediciones'
        sheet.merge_cells('W2:W3')

        sheet['X2'].border = borde_negro_grueso
        sheet['X2'].alignment = centrar_texto
        sheet['X2'].font = Font(bold=True)
        sheet['X2'] = 'FASE A'
        sheet.merge_cells('X2:Z2')


        sheet['AA2'].border = borde_negro_grueso
        sheet['AA2'].alignment = centrar_texto
        sheet['AA2'].font = Font(bold=True)
        sheet['AA2'] = 'FASE B'
        sheet.merge_cells('AA2:AC2')

        sheet['AA2'].border = borde_negro_grueso
        sheet['AA2'].alignment = centrar_texto
        sheet['AA2'].font = Font(bold=True)
        sheet['AA2'] = 'FASE B'
        sheet.merge_cells('AA2:AC2')


        #a partir de aqui agregamos los datos
        aux_init = 4
        
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['YEAR']
            sheet[f'B{aux_init}'] = i['MES']
            sheet[f'C{aux_init}'] = i['DIA']
            sheet[f'D{aux_init}'] = i['CODIGO']
            sheet[f'E{aux_init}'] = i['TIPO']
            sheet[f'F{aux_init}'] = i['SUBESTACION']
            sheet[f'G{aux_init}'] = i['ALIMENTADOR']
            sheet[f'H{aux_init}'] = i['NUM DE FASES']
            sheet[f'I{aux_init}'] = i['F-F']
            sheet[f'J{aux_init}'] = i['F-N']
           
            aux_init += 1


        workbook.save(final_path)

    #df = pd.DataFrame(muestra_generada)
    #df.to_excel(final_path, index=False)



